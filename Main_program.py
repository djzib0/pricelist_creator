import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from openpyxl import *
import os
from openpyxl.styles import *
from openpyxl.utils import get_column_letter


app_name = "Price-list creator"

# variables for widgets
file_name_pl = "Nazwa tworzonego pliku:"
amt_of_valid_files_pl = "Ilość plików do przetworzenia:"
amt_of_invalid_files_pl = "Ilość plików nie spełniających warunków"
app_language_pl = "Język"
file_name = 'Lista cen test.xlsx'

xlsx_files_list = []
valid_xlsx_files = []
invalid_xlsx_files = []
file_name = 'Lista cen test.xlsx'

# Creates list of excel files from which data will be collected

def select_xlsx_files():
    valid_files = []
    files = os.listdir()
    for file in files:
        if 'xlsx' in file and file != str(file_name):
            valid_files.append(file)
    print(valid_files)
    return valid_files

def select_valid_xlsx_files(file_list):
    import openpyxl
    """Sorts chosen xlsx files to valid or invalid."""
    list = file_list
    print("Kurwa! marker")
    print(list)
    valid_xlsx_files = []
    checklist = ['Customer', 'Ref. customer', 'Description', 'Drawing/ident']
    error_list = []

    for file in list:
        print("Nazwa pliku " + str(file))
        wb_to_read = openpyxl.load_workbook(str(file), data_only=True)
        ws_to_read = wb_to_read.active
        error_list = []
        for row_num in range(3, 7):
            if ws_to_read.cell(row=row_num, column=1).value in checklist and ws_to_read.cell(row=row_num, column=2).value != None:
                print("JEST OK")
            else:
                print("Error")
                error_list.append("Error")
        if len(error_list) == 0:
            valid_xlsx_files.append(file)
        else:
            invalid_xlsx_files.append(file)

    print("To jest kurwa mać poprawna lista xlsx")
    print(valid_xlsx_files)
    return valid_xlsx_files, invalid_xlsx_files


# write a code to create a file list
#xlsx_files_list = ['empty_book.xlsx', 'empty_book_1.xlsx']

def create_list_file_pl(files_list, file_name):
    import openpyxl
    list = files_list
    final_file_name = str(file_name) + ".xlsx"
    """Creates pricelist from excel files"""
    # creating top row with column names
    row = 1
    wb = Workbook()
    ws = wb.active
    column_names = ['Poz.', 'Nazwa', 'Rysunek', 'Cena/szt. [PLN]', 'Uwagi']

    # setting column widths
    ws.column_dimensions['A'].width = len(column_names[0]) + 1


    for i in range(1, len(column_names)+1):
        ws.cell(row=row, column=i).value = str(column_names[i-1])
        # Setting styles of top row with column names
        grey_fill = PatternFill(fill_type='solid',
                           start_color='00C0C0C0',
                           end_color='00C0C0C0')

        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal = 'center')
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'),
                            )

        # Formatting cells.
        ws.cell(row=row, column=i).fill = grey_fill
        ws.cell(row=row, column=i).font = bold_font
        ws.cell(row=row, column=i).alignment = center_alignment
        ws.cell(row=row, column=i).border = thin_border
    # Creating column names
    ### Dodać formatowanie!!!!!! Szerokość kolumn ustawić na podstawie najdłuższej nazwy
    #### Szerokość kolumny, czcionka, wielkość czcionki, ramkę tabeli, kolor tła,
    ##### Następny etap: jeżeli uwag jest więcej niż jedna, dodać każdą w nowym wierszu, ale scalić pozostałe i wyśrodkować!!!!
    ##### Zacząć od uwag???
    row = 2
    pos_number = 1
    name = ''
    drawing_number = ''
    price = ''
    remarks = ''
    for file in list:
        column_names = []
        wb_to_read = openpyxl.load_workbook(str(file), data_only=True)
        ws_to_read = wb_to_read.active
        for cell_row in range(1, 100):  # zmienić na max_row
            if ws_to_read.cell(row=cell_row, column=1).value == 'Description':
                name = ws_to_read.cell(row=cell_row, column=2).value
            if ws_to_read.cell(row=cell_row, column=1).value == 'Drawing/ident':
                drawing_number = ws_to_read.cell(row=cell_row, column=2).value
            if ws_to_read.cell(row=cell_row, column=6).value == 'Salesprice/piece':
                price = ws_to_read.cell(row=cell_row, column=7).value
        column_names.append(pos_number)
        column_names.append(name)
        column_names.append(drawing_number)
        column_names.append(price)
        column_names.append(remarks)
        for i in range(1, len(column_names)+1):
            ws.cell(row=row, column=i).value = str(column_names[i-1])
            ws.cell(row=row, column=i).alignment = center_alignment
            ws.cell(row=row, column=i).border = thin_border

        row += 1
        pos_number += 1

    # sets width of each column, based on longest cell data
    last_row = ws.max_row
    last_column = ws.max_column
    last_row = int(last_row)
    last_column = int(last_column)
    width_of_column = 0
    values = []

    for column in range(1, last_column + 1): # iteration through columns
        values = []  # temporary list for cells data from each column
        for row in range(1, last_row + 1): # iteration through rows
            values.append(ws.cell(row=row, column=column).value) # adding values to temporary list
        width_of_column = max(values, key=len) # from each column longest word is taken to determine width of column
        column_letter = get_column_letter(column) # getting letter of iterated column
        ws.column_dimensions[column_letter].width = len(width_of_column) + 1 # setting width of current column acc. longest word + 1

    # formatting cells in column D as float
    for row in range(2, last_row + 1): # iteration through cells
        print(ws.cell(row=row, column=4).value)
        ws.cell(row=row, column=4).value = float(ws.cell(row=row, column=4).value)
        ws.cell(row=row, column=4).number_format = '#,##0.00'

    wb.save(str(final_file_name))

class MainPage(QWidget):
    def __init__(self, title):
        super().__init__() # inherit init of QWidget
        self.title = title
        self.left = 250
        self.top = 250
        self.width = 600
        self.height = 400
        self.widget_pos_x = 10
        self.widget_pos_y = 50
        self.label_width = 180 # set height of labels
        self.label_height = 50 # sets height of labels
        self.valid_xlsx_files = []
        self.valid_files = []
        self.invalid_files = []
        self.file_name = ""
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        self.create_widgets_pl()

    def create_widgets_pl(self):
        # Widgets
        """Row 0"""
        self.file_name_lbl = QLabel(self, text=str("Lista cen"))
        self.file_name_lbl.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y,
                                             self.label_width, self.label_height))
        self.file_name_lbl.setWordWrap(True) # allow word-wrap
        self.widget_pos_y += 50

        self.amt_of_valid_files_lbl = QLabel(self, text=str(amt_of_valid_files_pl))
        self.amt_of_valid_files_lbl.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y,
                                             self.label_width, self.label_height))
        self.amt_of_valid_files_lbl.setWordWrap(True) # allow word-wrap
        self.widget_pos_y += 50

        self.amt_of_invalid_files_lbl = QLabel(self, text=str(amt_of_invalid_files_pl))
        self.amt_of_invalid_files_lbl.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y,
                                             self.label_width, self.label_height))
        self.amt_of_invalid_files_lbl.setWordWrap(True) # allow word-wrap
        self.widget_pos_y += 50

        self.create_list_btn = QPushButton(self, text="Generuj")
        self.create_list_btn.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y+20,
                                               80, 30))
        self.create_list_btn.clicked.connect(self.generate_file)

        """Row 1"""
        self.widget_pos_y = 0 # resets y coordinate for second row of Widgets

        self.app_language_lbl = QLabel(self, text=str(app_language_pl))
        self.app_language_lbl.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y,
                                                self.label_width, self.label_height))
        self.app_language_pl_radio = QRadioButton('POL', self)
        self.app_language_pl_radio.setChecked(True)
        self.app_language_pl_radio.move(self.widget_pos_x + 50, self.widget_pos_y+16.5)

        self.app_language_eng_radio = QRadioButton('ANG', self)
        self.app_language_eng_radio.move(self.widget_pos_x + 100, self.widget_pos_y+16.5)
        self.widget_pos_y += 65

        self.file_name_entry = QLineEdit(self, text="Lista cen")
        self.file_name_entry.setGeometry(QRect(self.widget_pos_x+200, self.widget_pos_y,
                                               100, 20))
        self.widget_pos_y += 35
        self.num_valid_files_lbl = QLabel(self, text=str(len(self.valid_files))) ### Tutaj odnieść się do listy plików
        self.num_valid_files_lbl.setGeometry(QRect(self.widget_pos_x+200, self.widget_pos_y,
                                                self.label_width, self.label_height))

        self.widget_pos_y += 45

        self.num_invalid_files_lbl = QLabel(self, text=str(len(self.invalid_files))) ### Tutaj odnieść się do listy plików
        self.num_invalid_files_lbl.setGeometry(QRect(self.widget_pos_x+200, self.widget_pos_y,
                                                self.label_width, self.label_height))

        self.show()

    @pyqtSlot()
    def create_file_name(self):
        file_name = self.file_name_entry.text()
        print("Nazwa pliku to KURWA MAĆ " + str(file_name))
        return file_name

    @pyqtSlot()
    def generate_file(self):
        file_name = self.create_file_name()
        self.valid_xlsx_files = select_xlsx_files()
        print("\nTo jest list wybranych plików z rozszerzeniem xlsx")
        print(self.valid_xlsx_files)
        self.valid_files, self.invalid_files = select_valid_xlsx_files(self.valid_xlsx_files)
        print("\nTo jest lista poprawnych plików xlsx")
        print(self.valid_files)
        print("\nTo jest lista niepoprawnych plików xlsx")
        print(self.invalid_files)
        create_list_file_pl(self.valid_files, file_name)


def main():
    app = QApplication(sys.argv)
    w = MainPage(app_name)
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
