import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from openpyxl import *
import os
from openpyxl.styles import *
from openpyxl.utils import get_column_letter


app_name = "Price-list creator"

# variables for widgets
file_name_pl = "Nazwa tworzonego pliku:"
amt_of_valid_files_pl = "Ilość plików do przetworzenia:"
amt_of_invalid_files_pl = "Ilość plików nie spełniających warunków"
app_language_pl = "Język"
app_currency_pl = "Waluta"
file_name = 'Lista cen test.xlsx'

xlsx_files_list = []
valid_xlsx_files = []
invalid_xlsx_files = []
file_name = 'Lista cen test.xlsx'

# Creates list of excel files from which data will be collected

def select_xlsx_files():
    valid_files = []
    files = os.listdir()
    for file in files:
        if 'xlsx' in file and file != str(file_name):
            valid_files.append(file)
    print(valid_files)
    return valid_files






class MainPage(QWidget):
    def __init__(self, title):
        super().__init__() # inherit init of QWidget
        import openpyxl
        self.title = title
        self.left = 250
        self.top = 250
        self.width = 600
        self.height = 400
        self.widget_pos_x = 10
        self.widget_pos_y = 20
        self.label_width = 180 # set height of labels
        self.label_height = 50 # sets height of labels
        self.valid_xlsx_files = []
        self.valid_files = []
        self.invalid_files = []
        self.file_name = ""
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        self.create_widgets_pl()

    def create_widgets_pl(self):
        # Widgets
        """Row 0"""
        self.file_name_lbl = QLabel(self, text=str("Nazwa pliku"))
        self.file_name_lbl.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y+30,
                                             self.label_width, self.label_height))
        self.file_name_lbl.setWordWrap(True) # allow word-wrap
        self.widget_pos_y += 115

        self.amt_of_valid_files_lbl = QLabel(self, text=str(amt_of_valid_files_pl))
        self.amt_of_valid_files_lbl.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y,
                                             self.label_width, self.label_height))
        self.amt_of_valid_files_lbl.setWordWrap(True) # allow word-wrap
        self.widget_pos_y += 50

        self.amt_of_invalid_files_lbl = QLabel(self, text=str(amt_of_invalid_files_pl))
        self.amt_of_invalid_files_lbl.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y,
                                             self.label_width, self.label_height))
        self.amt_of_invalid_files_lbl.setWordWrap(True) # allow word-wrap
        self.widget_pos_y += 50

        self.create_list_btn = QPushButton(self, text="Generuj")
        self.create_list_btn.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y+20,
                                               80, 30))
        self.create_list_btn.clicked.connect(self.generate_file)

        """Row 1"""
        self.widget_pos_y = 0 # resets y coordinate for second row of Widgets

        self.app_language_lbl = QLabel(self, text=str(app_language_pl))
        self.app_language_lbl.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y,
                                                self.label_width, self.label_height))

        self.app_currency_lbl = QLabel(self, text=str(app_currency_pl))
        self.app_currency_lbl.setGeometry(QRect(self.widget_pos_x, self.widget_pos_y+20,
                                                self.label_width, self.label_height))

        button_layout_1 = QVBoxLayout()
        self.app_language_pl_radio = QRadioButton('POL', self)
        self.app_language_pl_radio.move(self.widget_pos_x + 50, self.widget_pos_y+16.5)

        self.app_language_eng_radio = QRadioButton('ANG', self)
        self.app_language_eng_radio.move(self.widget_pos_x + 110, self.widget_pos_y+16.5)

        button_layout_1.addWidget(self.app_language_pl_radio)
        button_layout_1.addWidget(self.app_language_eng_radio)

        self.widget_pos_y += 35

        self.app_currency_pl_radio = QRadioButton('PLN', self)
        self.app_currency_pl_radio.move(self.widget_pos_x + 50, self.widget_pos_y)

        self.app_currency_eng_radio = QRadioButton('€', self)
        self.app_currency_eng_radio.move(self.widget_pos_x + 110, self.widget_pos_y)

        self.bg_1 = QButtonGroup() # Radiobutton group for language choice
        self.bg_1.addButton(self.app_language_pl_radio)
        self.bg_1.addButton(self.app_language_eng_radio)

        self.bg_2 = QButtonGroup() # Radiobutton group for currency choice
        self.bg_2.addButton(self.app_currency_pl_radio)
        self.bg_2.addButton(self.app_currency_eng_radio)
        self.app_currency_pl_radio.setChecked(True)
        self.app_language_pl_radio.setChecked(True)

        self.widget_pos_y += 65

        self.file_name_entry = QLineEdit(self, text="Lista cen")
        self.file_name_entry.setGeometry(QRect(self.widget_pos_x+200, self.widget_pos_y,
                                               100, 20))
        self.widget_pos_y += 35
        self.num_valid_files_lbl = QLabel(self, text=str(len(self.valid_files))) ### Tutaj odnieść się do listy plików
        self.num_valid_files_lbl.setGeometry(QRect(self.widget_pos_x+200, self.widget_pos_y,
                                                self.label_width, self.label_height))

        self.widget_pos_y += 45

        self.num_invalid_files_lbl = QLabel(self, text=str(len(self.invalid_files))) ### Tutaj odnieść się do listy plików
        self.num_invalid_files_lbl.setGeometry(QRect(self.widget_pos_x+200, self.widget_pos_y,
                                                self.label_width, self.label_height))

        self.show()

    def check_if_files_are_closed(self, files):
        import openpyxl
        """Checks if any file is opened."""
        try:
            for file in files:
                wb = openpyxl.load_workbook(str(file))
        except PermissionError:
            self.show_open_file_error()


    def check_language(self):
        """Sets language for column names in final file."""
        language = ''
        if self.app_language_pl_radio.isChecked():
            return 'pol'
        elif self.app_language_eng_radio.isChecked():
            return 'eng'

    def check_currency(self):
        currency = ''
        if self.app_currency_pl_radio.isChecked():
            return 'pln'
        elif self.app_currency_eng_radio.isChecked():
            return 'euro'

    """Message windows"""
    def show_file_exists_msg(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setText("Plik o takiej nazwiej już istnieje. Czy chcesz go nadpisać?")
        msg.setWindowTitle("Wystąpił błąd!")
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        retval = msg.exec_()
        #if retval == QMessageBox.Ok:
        #    print("OK Clicked")
        #elif retval == QMessageBox.Cancel:
        #    print("Cancel Clicked")
        return retval

    def show_open_file_error(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("Zamknij wszystkie pliki w katalogu z rozszerzeniem xlsx.")
        msg.setWindowTitle("Wystąpił błąd!")
        msg.setStandardButtons(QMessageBox.Ok)
        retval = msg.exec_()
        #if retval == QMessageBox.Ok:
        #    print("OK Clicked")
        #elif retval == QMessageBox.Cancel:
        #    print("Cancel Clicked")
        return retval

    def show_file_created_msg(self, file_name):
        file = file_name
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("Plik \"" + str(file) + ".xlsx\" został utworzony.")
        msg.setWindowTitle("Informacja")
        msg.setStandardButtons(QMessageBox.Ok)
        retval = msg.exec()

    """Main functions"""
    def create_list_file_pl(self, files_list, file_name, chosen_language, chosen_currency):
        """Creates final file with the given name, in selected language"""

        import openpyxl
        language = chosen_language
        currency = chosen_currency
        list = files_list
        final_file_name = str(file_name) + ".xlsx"
        """Creates pricelist from excel files"""
        # creating top row with column names
        row = 1
        wb = Workbook()
        ws = wb.active
        column_names_pl = ['Poz.', 'Nazwa', 'Rysunek', 'Cena/szt.', 'Uwagi']
        column_names_eng = ['Pos.', 'Name', 'Drawing', 'Price/pc.', 'Uwagi']
        if chosen_currency == 'pln':
            column_names_pl[3] += str(' [PLN]')
            column_names_eng[3] += str(' [PLN]')
        elif chosen_currency == 'euro':
            column_names_pl[3] += str(' [€]')
            column_names_eng[3] += str(' [€]')

        # setting column A width
        ws.column_dimensions['A'].width = len(column_names_pl[0]) + 1


        for i in range(1, len(column_names_pl)+1):
            if language == 'pol':
                ws.cell(row=row, column=i).value = str(column_names_pl[i-1])
            elif language == 'eng':
                ws.cell(row=row, column=i).value = str(column_names_eng[i-1])
            # Setting styles of top row with column names
            grey_fill = PatternFill(fill_type='solid',
                               start_color='00C0C0C0',
                               end_color='00C0C0C0')

            bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal = 'center')
            thin_border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'),
                                )

            # Formatting cells.
            ws.cell(row=row, column=i).fill = grey_fill
            ws.cell(row=row, column=i).font = bold_font
            ws.cell(row=row, column=i).alignment = center_alignment
            ws.cell(row=row, column=i).border = thin_border
        # Creating column names
        ##### Następny etap: jeżeli uwag jest więcej niż jedna, dodać każdą w nowym wierszu, ale scalić pozostałe i wyśrodkować!!!!
        ##### Zacząć od uwag???
        row = 2
        pos_number = 1
        name = ''
        drawing_number = ''
        price = ''
        remarks = ''
        for file in list:
            column_names = []
            wb_to_read = openpyxl.load_workbook(str(file), data_only=True)
            ws_to_read = wb_to_read.active
            for cell_row in range(1, 100):  # zmienić na max_row
                if ws_to_read.cell(row=cell_row, column=1).value == 'Description':
                    name = ws_to_read.cell(row=cell_row, column=2).value
                if ws_to_read.cell(row=cell_row, column=1).value == 'Drawing/ident':
                    drawing_number = ws_to_read.cell(row=cell_row, column=2).value
                if ws_to_read.cell(row=cell_row, column=6).value == 'Salesprice/piece':
                    price = ws_to_read.cell(row=cell_row, column=7).value
            column_names.append(pos_number)
            column_names.append(name)
            column_names.append(drawing_number)
            column_names.append(price)
            column_names.append(remarks)
            for i in range(1, len(column_names)+1):
                ws.cell(row=row, column=i).value = str(column_names[i-1])
                ws.cell(row=row, column=i).alignment = center_alignment
                ws.cell(row=row, column=i).border = thin_border

            row += 1
            pos_number += 1

        # sets width of each column, based on longest cell data
        last_row = ws.max_row
        last_column = ws.max_column
        last_row = int(last_row)
        last_column = int(last_column)
        width_of_column = 0
        values = []

        for column in range(1, last_column + 1): # iteration through columns
            values = []  # temporary list for cells data from each column
            for row in range(1, last_row + 1): # iteration through rows
                values.append(ws.cell(row=row, column=column).value) # adding values to temporary list
            width_of_column = max(values, key=len) # from each column longest word is taken to determine width of column
            column_letter = get_column_letter(column) # getting letter of iterated column
            ws.column_dimensions[column_letter].width = len(width_of_column) + 1 # setting width of current column acc. longest word + 1

        # formatting cells in column D as float
        for row in range(2, last_row + 1): # iteration through cells
            print(ws.cell(row=row, column=4).value)
            ws.cell(row=row, column=4).value = float(ws.cell(row=row, column=4).value)
            ws.cell(row=row, column=4).number_format = '#,##0.00'

        wb.save(str(final_file_name))

    def select_valid_xlsx_files(self, file_list):
        import openpyxl
        """Sorts chosen xlsx files to valid or invalid."""
        list = file_list
        print(list)

        valid_xlsx_files = []
        checklist = ['Customer', 'Ref. customer', 'Description', 'Drawing/ident']
        error_list = []

        for file in list:
            print("Nazwa pliku " + str(file))
            wb_to_read = openpyxl.load_workbook(str(file), data_only=True)
            ws_to_read = wb_to_read.active
            error_list = []
            for row_num in range(3, 7):
                if ws_to_read.cell(row=row_num, column=1).value in checklist and ws_to_read.cell(row=row_num, column=2).value != None:
                    print("JEST OK")
                else:
                    print("Error")
                    error_list.append("Error")
            if len(error_list) == 0:
                valid_xlsx_files.append(file)
            else:
                invalid_xlsx_files.append(file)

        print("To jest poprawna lista xlsx")
        print(valid_xlsx_files)
        return valid_xlsx_files, invalid_xlsx_files

    """Checking errors"""
    def check_if_files_are_closed(self, files):
        import openpyxl
        xlsx_files = files
        for file in xlsx_files:
            try:
                wb = openpyxl.load_workbook(file)
            except PermissionError:
                self.show_open_file_error()
                return True

    @pyqtSlot()
    def create_file_name(self):
        file_name = self.file_name_entry.text()
        return file_name

    @pyqtSlot()
    def generate_file(self):
        files_list = select_xlsx_files()
        opened_files = self.check_if_files_are_closed(files_list)
        if opened_files == None:
            file_name = self.create_file_name()
            language = self.check_language()
            currency = self.check_currency()
            if str(file_name) + ".xlsx" in files_list:
                choice = self.show_file_exists_msg()
                if choice == QMessageBox.Yes:
                    print("Yes Clicked")
                    self.valid_xlsx_files = select_xlsx_files()
                    print("\nTo jest lista wybranych plików z rozszerzeniem xlsx")
                    print(self.valid_xlsx_files)
                    self.valid_files, self.invalid_files = self.select_valid_xlsx_files(self.valid_xlsx_files)
                    print("\nTo jest lista poprawnych plików xlsx")
                    print(self.valid_files)
                    print("\nTo jest lista niepoprawnych plików xlsx")
                    print(self.invalid_files)
                    self.create_list_file_pl(self.valid_files, file_name, language, currency)
                    self.show_file_created_msg(file_name)
                elif choice == QMessageBox.No:
                    print("No Clicked")
            else:
                self.valid_xlsx_files = select_xlsx_files()
                self.valid_files, self.invalid_files = self.select_valid_xlsx_files(self.valid_xlsx_files)
                print(self.valid_files)
                self.create_list_file_pl(self.valid_files, file_name, language, currency)
                self.show_file_created_msg(file_name)


def main():
    app = QApplication(sys.argv)
    w = MainPage(app_name)
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
